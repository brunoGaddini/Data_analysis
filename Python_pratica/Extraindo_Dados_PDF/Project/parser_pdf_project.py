# Importing libraries
import openpyxl as xl
import pdfplumber
import os

# Searching the pdf files
for arquivo in os.listdir('C:\Python\EmpowerData\Python_pratica\Extraindo_Dados_PDF'):

    if arquivo.lower().endswith('.pdf'):
        try:
            # Opening excel file
            excel = xl.load_workbook(
                'C:\Python\EmpowerData\Python_pratica\Extraindo_Dados_PDF\Base de Dados Inspeções.xlsx')
            aba = excel.active
            linha_inicio = len(aba['A']) + 1

            # Read the pdf file and extract the data
            pdf = pdfplumber.open(f'C:\Python\\EmpowerData\\Python_pratica\\Extraindo_Dados_PDF\\{arquivo}')
            pagina = pdf.pages[0]

            # Entering data into the variable
            dados = pagina.extract_table()
            for indice, dados in enumerate(dados[1:], start=linha_inicio):
                if dados[0] == '':
                    pass

                else:
                    aba.cell(row=indice, column=1).value = dados[0]
                    aba.cell(row=indice, column=2).value = dados[1]
                    aba.cell(row=indice, column=3).value = dados[2]
                    aba.cell(row=indice, column=4).value = dados[3]
                    aba.cell(row=indice, column=5).value = dados[4]

            pdf.close()
            excel.save('C:\Python\EmpowerData\Python_pratica\Extraindo_Dados_PDF\Base de Dados Inspeções.xlsx')
            excel.close()

        except Exception as e:
            with open('log_erros.txt', 'a') as log:
                log.write(f'Ocorreu um erro ao extrair dados do arquivo {arquivo}.\n')
                log.write(f'Erro: {e}')

    else:
        with open('log_erros.txt', 'a') as log:
            log.write(f'O arquivo {arquivo} não é válido!\n')